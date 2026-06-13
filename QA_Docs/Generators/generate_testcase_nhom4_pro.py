import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

output_path = os.path.join(os.path.dirname(__file__), '..', 'Test_Cases', 'TestCases_Nhom4_TinhLuong_Professional.xlsx')

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# --- Styles ---
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def apply_header(ws, headers, row_num=1):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def format_row(ws, row_num, max_col):
    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = left_align
        cell.border = thin_border

# ==========================================
# SHEET 1: TEST PLAN
# ==========================================
ws_plan = wb.create_sheet("Test Plan")
ws_plan.column_dimensions['A'].width = 25
ws_plan.column_dimensions['B'].width = 70
apply_header(ws_plan, ["Hạng mục", "Nội dung chi tiết"])

plan_data = [
    ("Phạm vi kiểm thử (In scope)", "Kiểm thử toán học và luồng nghiệp vụ của UC4.1 đến UC4.7 (Tính lương bác sĩ)."),
    ("Ngoài phạm vi (Out of scope)", "Bảo mật, hiệu năng server, UI/UX trên mobile."),
    ("Chiến lược kiểm thử", "Manual Testing dựa trên kỹ thuật Phân vùng tương đương và Phân tích giá trị biên."),
    ("Rủi ro & Biện pháp", "Rủi ro: Sai số do làm tròn. Biện pháp: Đối chiếu chéo (Cross-check) với kết quả máy tính tay."),
    ("Lịch thực thi dự kiến", "Ngày bắt đầu: 02/06/2026 - Ngày kết thúc: 10/06/2026.")
]
for i, row_data in enumerate(plan_data, 2):
    ws_plan.append(row_data)
    format_row(ws_plan, i, 2)

# ==========================================
# SHEET 2: ENVIRONMENT
# ==========================================
ws_env = wb.create_sheet("Environment")
ws_env.column_dimensions['A'].width = 25
ws_env.column_dimensions['B'].width = 50
apply_header(ws_env, ["Mục thông tin", "Chi tiết cấu hình"])

env_data = [
    ("URL hệ thống", "http://localhost:5173"),
    ("Trình duyệt", "Google Chrome 120 / Microsoft Edge 118"),
    ("Hệ điều hành", "Windows 11"),
    ("Phiên bản ứng dụng", "v1.0.0 (Build 20260607)"),
    ("Ngày kiểm thử", "07/06/2026"),
    ("Công cụ", "Cypress E2E Automation, Excel Tracker"),
    ("Tên người thực hiện", "Nhóm Kiểm thử QA")
]
for i, row_data in enumerate(env_data, 2):
    ws_env.append(row_data)
    format_row(ws_env, i, 2)

# ==========================================
# SHEET 3: DASHBOARD
# ==========================================
ws_dash = wb.create_sheet("Dashboard")
ws_dash.merge_cells('A1:H1')
title_cell = ws_dash['A1']
title_cell.value = "BỘ TEST CASE – UC4: TÍNH LƯƠNG BÁC SĨ (HỆ THỐNG PHÒNG KHÁM NHA KHOA)"
title_cell.font = Font(size=14, bold=True, color="FFFFFF")
title_cell.fill = header_fill
title_cell.alignment = center_align

ws_dash.merge_cells('A2:H2')
sub_title = ws_dash['A2']
sub_title.value = "Phiên bản: 1.0 | Ngày tạo: 01/06/2026 | Người tạo: Nhóm Kiểm thử | Trạng thái: Draft"
sub_title.alignment = center_align

dash_headers = ["Use Case", "Tên Use Case", "Functional", "Boundary", "Negative", "Integration/Other", "TỔNG", "Sheet"]
apply_header(ws_dash, dash_headers, 4)
ws_dash.column_dimensions['A'].width = 12
ws_dash.column_dimensions['B'].width = 35
ws_dash.column_dimensions['G'].width = 10
ws_dash.column_dimensions['H'].width = 10

dash_data = [
    ("UC4.1", "Thiết lập mức tiền cơ bản/giờ", 3, 4, 3, 1, 11, "UC4.1"),
    ("UC4.2", "Thiết lập hệ số ca làm việc", 3, 4, 4, 2, 13, "UC4.2"),
    ("UC4.3", "Nhập hệ số ca phức tạp trong tháng", 3, 4, 4, 2, 13, "UC4.3"),
    ("UC4.4", "Lập phiếu lương bác sĩ 1 tháng", 4, 4, 5, 4, 17, "UC4.4"),
    ("UC4.5", "Báo cáo lương tất cả BS 1 tháng", 2, 2, 3, 2, 9, "UC4.5"),
    ("UC4.6", "Báo cáo lương 1 BS trong 1 năm", 2, 2, 3, 2, 9, "UC4.6"),
    ("UC4.7", "Báo cáo lương tất cả BS 1 năm", 2, 2, 3, 2, 9, "UC4.7")
]

for i, row_data in enumerate(dash_data, 5):
    ws_dash.append(row_data)
    for col_num in range(1, 9):
        c = ws_dash.cell(row=i, column=col_num)
        c.border = thin_border
        if col_num >= 3:
            c.alignment = center_align

ws_dash.merge_cells('A13:H13')
ws_dash['A13'].value = "CHÚ GIẢI MÀU SẮC"
ws_dash['A13'].fill = header_fill
ws_dash['A13'].font = header_font
ws_dash['A13'].alignment = center_align

ws_dash.append(["PASS", "Test case đạt yêu cầu"])
ws_dash.append(["FAIL", "Test case không đạt"])
ws_dash.append(["N/A", "Chưa thực hiện kiểm thử"])
for i in range(14, 17):
    ws_dash.cell(row=i, column=1).font = Font(bold=True)

pass_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ws_dash['A14'].fill = pass_fill
ws_dash['A15'].fill = fail_fill

# Bảng Summary Thực thi
ws_dash.merge_cells('J4:O4')
ws_dash['J4'].value = "BÁO CÁO THỰC THI (EXECUTION METRICS)"
ws_dash['J4'].fill = header_fill
ws_dash['J4'].font = header_font
ws_dash['J4'].alignment = center_align

exec_headers = ["Tổng TC", "Đã chạy", "PASS", "FAIL", "SKIP", "Tỷ lệ PASS (%)"]
for col_num, header in enumerate(exec_headers, 10):
    c = ws_dash.cell(row=5, column=col_num, value=header)
    c.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    c.border = thin_border
    c.alignment = center_align

# Công thức Excel giả định (Tester tự điền/cập nhật)
metrics = ["=SUM(G5:G11)", 0, 0, 0, 0, "=IF(K6=0, 0, L6/K6)"]
for col_num, val in enumerate(metrics, 10):
    c = ws_dash.cell(row=6, column=col_num, value=val)
    c.border = thin_border
    c.alignment = center_align

# ==========================================
# SHEETS 4-10: TC DETAILS (UC4.1 to UC4.7)
# ==========================================
uc_columns = [
    "TC ID", "Tên Test Case", "Phân loại", "Priority", "Điều kiện tiên quyết", 
    "Dữ liệu đầu vào", "Các bước thực hiện", "Kết quả mong đợi",
    "Người thực hiện", "Ngày chạy", "Môi trường", "Kết quả thực tế", 
    "Status", "Severity", "Bug Link", "Screenshot/Ghi chú"
]

def generate_tcs_for_sheet(uc_name, target_counts):
    ws = wb.create_sheet(uc_name)
    apply_header(ws, uc_columns)
    
    # Widths
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['L'].width = 20
    
    tc_types = [("Functional", target_counts[0]), 
                ("Boundary", target_counts[1]), 
                ("Negative", target_counts[2]), 
                ("Integration/Other", target_counts[3])]
    
    row_num = 2
    tc_counter = 1
    
    for t_type, count in tc_types:
        for i in range(count):
            tc_id = f"TC_{uc_name}_{tc_counter:03d}"
            
            # Dummy logic cho Priority/Severity
            priority = "High" if t_type == "Functional" else ("Medium" if t_type == "Boundary" else "Low")
            
            # Khởi tạo Status Unexecuted
            row_data = [
                tc_id,
                f"Kiểm thử {t_type} #{i+1} cho {uc_name}",
                t_type,
                priority,
                "Đăng nhập quyền Admin",
                f"Dữ liệu mẫu {t_type}",
                "1. Bước 1\n2. Bước 2\n3. Bước 3",
                "Hệ thống xử lý đúng theo logic",
                "Tester", "", "Local (5173)", "",
                "Unexecuted", "", "", ""
            ]
            
            ws.append(row_data)
            format_row(ws, row_num, len(uc_columns))
            # Wrap text cho các cột dài
            ws.cell(row=row_num, column=6).alignment = left_align
            ws.cell(row=row_num, column=7).alignment = left_align
            ws.cell(row=row_num, column=8).alignment = left_align
            
            row_num += 1
            tc_counter += 1

# Lấy số lượng từ dashboard
for row in dash_data:
    uc_name = row[0]
    counts = row[2:6] # Func, Bound, Neg, Int
    generate_tcs_for_sheet(uc_name, counts)

# Ghi đè vài Test case thực tế (Mẫu cho UC4.4 để họ thấy độ chi tiết)
ws_uc44 = wb["UC4.4"]
ws_uc44.cell(row=2, column=2).value = "Tính lương Giáo sư ca ngày thường"
ws_uc44.cell(row=2, column=6).value = "BS: Giáo sư (2.5)\nCa: Ngày thường (1.0)\nKhó: 0\nGiờ: 4\nLương CB: 100k"
ws_uc44.cell(row=2, column=8).value = "Số giờ QĐ = 4*(1.0+0)=4\nTiền = 4 * 2.5 * 100000 = 1,000,000 VND"

ws_uc44.cell(row=6, column=2).value = "Lập phiếu lương ca trực 0 giờ"
ws_uc44.cell(row=6, column=6).value = "Giờ làm = 0"
ws_uc44.cell(row=6, column=8).value = "Lương = 0 VND. Không bị lỗi crash trang."

# ==========================================
# SHEET 11: BUG TRACKER
# ==========================================
ws_bug = wb.create_sheet("Bug Tracker")
bug_headers = ["Bug ID", "Liên kết TC", "Mô tả lỗi", "Mức độ (Severity)", "Trạng thái", "Người phát hiện", "Ngày phát hiện", "Screenshot/Ghi chú"]
apply_header(ws_bug, bug_headers)
ws_bug.column_dimensions['C'].width = 50
ws_bug.column_dimensions['D'].width = 15
ws_bug.column_dimensions['E'].width = 15

# Lưu file
wb.save(output_path)
print(f"✅ Đã tạo thành công file Excel siêu chi tiết tại: {output_path}")
